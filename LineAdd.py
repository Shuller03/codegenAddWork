from openpyxl import load_workbook
from unittest import TestCase
import glob,os
import io

class LineAdd(TestCase):

    os.chdir('cases')
    fileTypes = []

    def test_fileTypes(self, types = fileTypes):

        for file in glob.glob('DU*'):
            l, r = file.split('_')
            l = l[2] + l[3]
            types.append(l.upper())
        print(types)
        return types


    def test_getColumn(self ,fileTypes = ['CD', 'IL', 'IR', 'OL', 'OR']):

        sheet = load_workbook('FVB_Connections_test.xlsx')['Лист1']
        typesNavi = {}
        sheetSlice = [sheet['A2'].value, sheet['B2'].value,
                      sheet['C2'].value, sheet['D2'].value,
                      sheet['E2'].value, sheet['F2'].value,
                      sheet['G2'].value, sheet['H2'].value]

        for element in sheetSlice:
            if element in fileTypes:
                typesNavi[chr(sheetSlice.index(element) + 65)] = element
        print(typesNavi)
        return typesNavi


    def test_getColumnElements(self, typesNavi = {'D': 'OL', 'E': 'IL', 'F': 'CD', 'G': 'IR', 'H': 'OR'}):

        sheet = load_workbook('FVB_Connections_test.xlsx')['Лист1']
        buff  =  []

        for k, v in typesNavi.items():      #сначала открытие файла, потом всё остальное (можно открытие файлов из списка)
            for i in range(3, 130):
                buff.append(sheet[k + str(i)].value)
            with io.open('DUCD_DA', encoding='utf-8') as file:
                for line in file:
                    c = 0
                    for i in range(0, buff.__len__()):

                        if str(buff[i]) in line:
                            c+=1
                            print(line)
                    print(c)